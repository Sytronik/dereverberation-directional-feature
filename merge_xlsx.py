import openpyxl
import numpy as np
import os
from os.path import join as pathjoin
from copy import copy
from argparse import ArgumentParser

parser = ArgumentParser()
parser.add_argument('path', nargs='?', type=str, const='./')
# parser.add_argument('--tag', nargs=1, type=str, default=('',))
# parser.add_argument('--delete', '-d', action='store_true')
ARGS = parser.parse_args()
files = [f.path
         for f in os.scandir(ARGS.path)
         if f.name.endswith('.xlsx')]
if not files:
    print('No files.')
    exit()

kind = dict()
merged_cells = dict()
for fname in files:
    fname_split = fname.split('_')
    if len(fname_split) == 1:
        continue
    key = fname_split[1]
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    if key not in kind:
        kind[key] = [ws['A:A']]

    if key not in merged_cells:
        merged_cells[key] = copy(ws.merged_cells.ranges)
    else:
        ranges = copy(ws.merged_cells.ranges)
        for range_ in ranges:
            range_.max_col += len(kind[key]) - 1
            range_.min_col += len(kind[key]) - 1
        merged_cells[key].extend(ranges)
    kind[key].extend(list(ws['B:E']))

    print(os.path.basename(fname))
print()

wb = openpyxl.Workbook()
wb.remove(wb.active)
for idx, (key, value) in enumerate(kind.items()):
    ws = wb.create_sheet(key, idx)
    for idx_col, col in enumerate(value):
        for idx_row, cell in enumerate(col):
            ws.cell(idx_row+1, idx_col+1).value = cell.value
    for range_ in merged_cells[key]:
        ws.merge_cells(str(range_))


fname_result = pathjoin(ARGS.path, f'merged.xlsx')
while True:
    try:
        wb.save(fname_result)
        print(fname_result)
        break
    except PermissionError:
        fname_result = fname_result.replace('.xlsx', '_.xlsx')
