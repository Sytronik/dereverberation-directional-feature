import openpyxl
import numpy as np
import json
import os
from os.path import join as pathjoin
from collections import defaultdict
from argparse import ArgumentParser

parser = ArgumentParser()
# parser.add_argument('title2', nargs=1, type=str)
# parser.add_argument('--title1', nargs=1, type=str, metavar='title1')
parser.add_argument('path', nargs='?', type=str, const='./')
parser.add_argument('--tag', nargs=1, type=str, default=('',))
parser.add_argument('--delete', '-d', action='store_true')
ARGS = parser.parse_args()
files = [f.path
         for f in os.scandir(ARGS.path)
         if f.name.endswith('.json')]
if not files:
    print('No files.')
    exit()

title = ARGS.tag[0]
col_title = ''

common = defaultdict(lambda: 0)
files_chosen = []
for fname in files:
    fname_split = os.path.basename(fname)
    fname_split = fname_split.replace('-tag-', '_')
    fname_split = fname_split.replace('.json', '').split('_')
    if title not in fname_split:
        continue

    files_chosen.append(fname)
    print(fname)

    if not col_title:
        col_title = fname_split[2]

    for item in fname_split:
        if '.' not in item:
            common[item] += 1

if not col_title:
    print('No files.')
    exit()

print()

col_names = [k for k, v in common.items() if v == 1]
step = None
measures = dict()

for fname, col_name in zip(files_chosen, col_names):
    with open(fname) as f:
        data = json.loads(f.read())
        data = np.array(data)
    if ARGS.delete:
        os.remove(fname)
    if not step:  # and ARGS.title1:
        step = data[:, 1].tolist()
    measures[col_name] = data[:, 2]


wb = openpyxl.Workbook()
ws = wb.active
# cols = ['B', 'C', 'D']  # if step else ['A', 'B', 'C']
cols = [chr(ord('B')+i) for i in range(len(measures))]

ws['A1'] = title
ws[f'{cols[0]}1'] = col_title
ws.merge_cells(f'{cols[0]}1:{cols[-1]}1')

ws['A2'] = 'Step' if step else ''
for idx, (cell,) in enumerate(ws[f'A3:A{3 + len(step) - 1}']):
    cell.value = step[idx]

for ii, (key, value) in enumerate(measures.items()):
    ws[f'{cols[ii]}2'] = key
    for jj, (cell,) in enumerate(ws[f'{cols[ii]}3:{cols[ii]}{3 + value.shape[0] - 1}']):
        cell.value = value[jj]

fname_result = pathjoin(ARGS.path, f'merged_{title}_{col_title}.xlsx')
while True:
    try:
        wb.save(fname_result)
        print(fname_result)
        break
    except PermissionError:
        fname_result = fname_result.replace('.xlsx', '_.xlsx')
